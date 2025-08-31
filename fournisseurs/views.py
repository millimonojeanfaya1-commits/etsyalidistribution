from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.core.paginator import Paginator
from django.http import HttpResponse
from .models import Fournisseur, Produit, Livraison
from .forms import FournisseurForm, ProduitForm, LivraisonForm


@login_required
def fournisseur_list(request):
    """
    Liste des fournisseurs avec recherche
    """
    fournisseurs = Fournisseur.objects.all()
    
    # Recherche
    search = request.GET.get('search')
    if search:
        fournisseurs = fournisseurs.filter(
            Q(nom__icontains=search) | 
            Q(telephone__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(fournisseurs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'title': 'Liste des Fournisseurs',
        'page_obj': page_obj,
        'search': search,
    }
    return render(request, 'fournisseurs/fournisseur_list.html', context)


@login_required
def fournisseur_create(request):
    """
    Créer un nouveau fournisseur avec ses produits fournis
    """
    if request.method == 'POST':
        try:
            import json
            
            # Récupérer les données du fournisseur
            nom = request.POST.get('nom')
            telephone = request.POST.get('telephone')
            email = request.POST.get('email')
            adresse = request.POST.get('adresse')
            
            if not nom:
                messages.error(request, 'Le nom du fournisseur est obligatoire.')
                return render(request, 'fournisseurs/fournisseur_form.html', {
                    'title': 'Nouveau Fournisseur',
                    'produits': Produit.objects.all(),
                })
            
            # Créer le fournisseur
            fournisseur = Fournisseur.objects.create(
                nom=nom,
                telephone=telephone or '',
                email=email or '',
                adresse=adresse or ''
            )
            
            # Traiter les produits fournis
            produits_count = 0
            for key, value in request.POST.items():
                if key.startswith('produits[') and key.endswith(']'):
                    try:
                        produit_data = json.loads(value)
                        
                        # Récupérer le produit
                        produit = get_object_or_404(Produit, id=produit_data['id'])
                        
                        # Créer une livraison pour ce produit
                        numero_livraison = f"LIV-{fournisseur.id}-{produits_count + 1:03d}"
                        
                        Livraison.objects.create(
                            numero_enregistrement=numero_livraison,
                            date='2025-08-15',  # Date du jour
                            fournisseur=fournisseur,
                            produit=produit,
                            quantite_livree=float(produit_data['quantite']),
                            prix_achat_unitaire=float(produit_data['prix_unitaire']),
                            observations=produit_data.get('observations', '')
                        )
                        produits_count += 1
                        
                    except (json.JSONDecodeError, KeyError, ValueError) as e:
                        messages.warning(request, f'Erreur lors du traitement d\'un produit: {str(e)}')
                        continue
            
            if produits_count > 0:
                messages.success(request, f'Fournisseur "{nom}" créé avec succès avec {produits_count} produit(s) fourni(s)!')
            else:
                messages.success(request, f'Fournisseur "{nom}" créé avec succès!')
                
            return redirect('fournisseurs:fournisseur_list')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création: {str(e)}')
    
    context = {
        'title': 'Nouveau Fournisseur',
        'produits': Produit.objects.all(),
    }
    return render(request, 'fournisseurs/fournisseur_form.html', context)


@login_required
def fournisseur_edit(request, pk):
    """
    Modifier un fournisseur
    """
    fournisseur = get_object_or_404(Fournisseur, pk=pk)
    
    if request.method == 'POST':
        form = FournisseurForm(request.POST, instance=fournisseur)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fournisseur modifié avec succès!')
            return redirect('fournisseurs:fournisseur_list')
    else:
        form = FournisseurForm(instance=fournisseur)
    
    context = {
        'title': f'Modifier {fournisseur.nom}',
        'form': form,
        'fournisseur': fournisseur,
    }
    return render(request, 'fournisseurs/fournisseur_form.html', context)


@login_required
def livraison_list(request):
    """
    Liste des livraisons avec filtres
    """
    livraisons = Livraison.objects.select_related('fournisseur', 'produit').all()
    
    # Filtres
    fournisseur_id = request.GET.get('fournisseur')
    if fournisseur_id:
        livraisons = livraisons.filter(fournisseur_id=fournisseur_id)
    
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    if date_debut:
        livraisons = livraisons.filter(date__gte=date_debut)
    if date_fin:
        livraisons = livraisons.filter(date__lte=date_fin)
    
    # Pagination
    paginator = Paginator(livraisons, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistiques
    total_achats = livraisons.aggregate(Sum('montant_total_achat'))['montant_total_achat__sum'] or 0
    
    context = {
        'title': 'Historique des Livraisons',
        'page_obj': page_obj,
        'fournisseurs': Fournisseur.objects.all(),
        'total_achats': total_achats,
        'filters': {
            'fournisseur': fournisseur_id,
            'date_debut': date_debut,
            'date_fin': date_fin,
        }
    }
    return render(request, 'fournisseurs/livraison_list.html', context)


@login_required
def livraison_export_excel(request):
    """Exporter la liste des livraisons en Excel (en respectant les filtres)"""
    livraisons = Livraison.objects.select_related('fournisseur', 'produit').all()

    # Filtres
    fournisseur_id = request.GET.get('fournisseur')
    if fournisseur_id:
        livraisons = livraisons.filter(fournisseur_id=fournisseur_id)

    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    if date_debut:
        livraisons = livraisons.filter(date__gte=date_debut)
    if date_fin:
        livraisons = livraisons.filter(date__lte=date_fin)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Livraisons'
    ws.append([
        'N° Enregistrement', 'Date', 'Fournisseur', 'Produit',
        'Quantité livrée', "Unité", "Prix d'achat unitaire", 'Montant total achat',
        'Observations'
    ])

    for l in livraisons.order_by('-date'):
        prix_unit = float(l.prix_achat_unitaire) if l.prix_achat_unitaire is not None else 0
        montant_total = float(l.montant_total_achat) if l.montant_total_achat is not None else 0
        ws.append([
            l.numero_enregistrement,
            l.date.strftime('%d/%m/%Y') if l.date else '',
            l.fournisseur.nom if l.fournisseur else '',
            l.produit.nom if l.produit else '',
            float(l.quantite_livree) if l.quantite_livree is not None else 0,
            getattr(l.produit, 'unite_mesure', '') or '',
            int(round(prix_unit)),
            int(round(montant_total)),
            l.observations or '',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="livraisons.xlsx"'
    wb.save(response)
    return response


@login_required
def livraison_create(request):
    """
    Enregistrer une nouvelle livraison
    """
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero_enregistrement = request.POST.get('numero_enregistrement')
            date = request.POST.get('date')
            fournisseur_id = request.POST.get('fournisseur')
            produit_id = request.POST.get('produit')
            quantite_livree = request.POST.get('quantite_livree')
            prix_achat_unitaire = request.POST.get('prix_achat_unitaire')
            observations = request.POST.get('observations')
            
            # Gérer les nouvelles entrées créées dynamiquement
            fournisseur = None
            if fournisseur_id and fournisseur_id.startswith('new_fournisseur_'):
                # Créer un nouveau fournisseur
                fournisseur = Fournisseur.objects.create(
                    nom="Nouveau Fournisseur",
                    adresse="Adresse à compléter",
                    telephone="",
                    email=""
                )
            else:
                fournisseur = get_object_or_404(Fournisseur, id=fournisseur_id) if fournisseur_id else None
            
            produit = None
            if produit_id and produit_id.startswith('new_product_'):
                # Créer un nouveau produit
                produit = Produit.objects.create(
                    nom="Nouveau Produit",
                    description="Produit créé depuis le formulaire de livraison",
                    unite_mesure="pièce",
                    prix_vente_conseille=0
                )
            else:
                produit = get_object_or_404(Produit, id=produit_id) if produit_id else None
            
            # Créer la livraison
            if fournisseur and produit:
                livraison = Livraison.objects.create(
                    numero_enregistrement=numero_enregistrement,
                    date=date,
                    fournisseur=fournisseur,
                    produit=produit,
                    quantite_livree=float(quantite_livree),
                    prix_achat_unitaire=float(prix_achat_unitaire),
                    observations=observations
                )
                messages.success(request, f'Livraison {numero_enregistrement} enregistrée avec succès!')
                return redirect('fournisseurs:livraison_list')
            else:
                messages.error(request, 'Erreur: Tous les champs obligatoires doivent être remplis.')
                
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'enregistrement: {str(e)}')
    
    # Préparer les données pour le template
    context = {
        'title': 'Nouvelle Livraison',
        'fournisseurs': Fournisseur.objects.all(),
        'produits': Produit.objects.all(),
        'today': '2025-08-15',  # Date du jour
    }
    return render(request, 'fournisseurs/livraison_form.html', context)


@login_required
def statistiques_fournisseurs(request):
    """
    Statistiques des achats par fournisseur
    """
    # Montant total par fournisseur
    stats_fournisseurs = Fournisseur.objects.annotate(
        total_achats=Sum('livraison__montant_total_achat'),
        nb_livraisons=Count('livraison')
    ).filter(total_achats__isnull=False).order_by('-total_achats')
    
    # Statistiques mensuelles (derniers 12 mois)
    from datetime import datetime, timedelta
    from django.utils import timezone
    from django.db.models.functions import TruncMonth
    
    date_limite = timezone.now().date() - timedelta(days=365)
    stats_mensuelles = Livraison.objects.filter(date__gte=date_limite).annotate(
        mois=TruncMonth('date')
    ).values('mois').annotate(
        total=Sum('montant_total_achat'),
        nb_livraisons=Count('id')
    ).order_by('mois')
    
    context = {
        'title': 'Statistiques Fournisseurs',
        'stats_fournisseurs': stats_fournisseurs,
        'stats_mensuelles': stats_mensuelles,
    }
    return render(request, 'fournisseurs/statistiques.html', context)


@login_required
def fournisseur_export_excel(request):
    """Exporter la liste des fournisseurs (avec recherche) en Excel"""
    qs = Fournisseur.objects.all().order_by('nom')
    search = request.GET.get('search')
    if search:
        qs = qs.filter(
            Q(nom__icontains=search) |
            Q(telephone__icontains=search) |
            Q(email__icontains=search)
        )

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Fournisseurs'
    ws.append(['Nom', 'Téléphone', 'Email', 'Adresse', "Date d'enregistrement"]) 
    for f in qs:
        ws.append([
            f.nom,
            f.telephone or '',
            f.email or '',
            f.adresse or '',
            f.date_creation.strftime('%d/%m/%Y') if f.date_creation else '',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="fournisseurs.xlsx"'
    wb.save(response)
    return response


@login_required
def produit_export_excel(request):
    """Exporter la liste des produits en Excel (avec recherche basique)"""
    produits = Produit.objects.all().order_by('nom')
    search = request.GET.get('search', '').strip()
    if search:
        produits = produits.filter(Q(nom__icontains=search) | Q(unite_mesure__icontains=search))

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Produits'
    ws.append(['Nom', 'Unité', 'Prix conseillé', 'Description', "Date d'enregistrement"]) 
    for p in produits:
        prix_conseille = float(p.prix_vente_conseille) if p.prix_vente_conseille is not None else 0
        ws.append([
            p.nom,
            p.unite_mesure,
            int(round(prix_conseille)),
            p.description or '',
            p.date_creation.strftime('%d/%m/%Y') if p.date_creation else '',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="produits.xlsx"'
    wb.save(response)
    return response
